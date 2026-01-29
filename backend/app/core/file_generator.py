"""
File Generator Module

Generates CSV and Excel files from extracted email data.
Creates structured, formatted output files.

Author: Email Extraction System
"""

import csv
import os
from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class CSVGenerator:
    """
    Generates CSV files from email data.
    """
    
    def __init__(self, 
                 delimiter: str = ',',
                 quoting: int = csv.QUOTE_MINIMAL,
                 encoding: str = 'utf-8'):
        """
        Initialize CSV generator.
        
        Args:
            delimiter: CSV delimiter character
            quoting: CSV quoting mode
            encoding: File encoding
        """
        self.delimiter = delimiter
        self.quoting = quoting
        self.encoding = encoding
    
    def generate(self, 
                emails: List[Dict[str, any]], 
                output_path: str,
                columns: Optional[List[str]] = None) -> str:
        """
        Generate CSV file from email data.
        
        Args:
            emails: List of email dictionaries
            output_path: Output file path
            columns: Column names to include (uses all if None)
            
        Returns:
            Path to generated file
        """
        if not emails:
            logger.warning("No emails provided for CSV generation")
            return None
        
        try:
            # Determine columns
            if columns is None:
                columns = list(emails[0].keys())
            
            # Write CSV file
            with open(output_path, 'w', newline='', encoding=self.encoding) as f:
                writer = csv.DictWriter(
                    f, 
                    fieldnames=columns,
                    delimiter=self.delimiter,
                    quoting=self.quoting,
                    extrasaction='ignore'  # Ignore extra fields
                )
                
                # Write header
                writer.writeheader()
                
                # Write data rows
                writer.writerows(emails)
            
            logger.info(f"Generated CSV file with {len(emails)} rows at {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate CSV: {str(e)}")
            raise
    
    def generate_simple(self, 
                       emails: List[str], 
                       output_path: str,
                       include_index: bool = True) -> str:
        """
        Generate simple CSV with just email addresses.
        
        Args:
            emails: List of email addresses
            output_path: Output file path
            include_index: Include row numbers
            
        Returns:
            Path to generated file
        """
        try:
            with open(output_path, 'w', newline='', encoding=self.encoding) as f:
                writer = csv.writer(f, delimiter=self.delimiter, quoting=self.quoting)
                
                # Write header
                if include_index:
                    writer.writerow(['Index', 'Email'])
                else:
                    writer.writerow(['Email'])
                
                # Write emails
                for idx, email in enumerate(emails, start=1):
                    if include_index:
                        writer.writerow([idx, email])
                    else:
                        writer.writerow([email])
            
            logger.info(f"Generated simple CSV with {len(emails)} emails at {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate simple CSV: {str(e)}")
            raise


class ExcelGenerator:
    """
    Generates Excel (.xlsx) files from email data with formatting.
    """
    
    # Default style configuration
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    CELL_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, 
                 apply_formatting: bool = True,
                 auto_filter: bool = True,
                 freeze_header: bool = True):
        """
        Initialize Excel generator.
        
        Args:
            apply_formatting: Apply styling to the worksheet
            auto_filter: Enable auto-filter on headers
            freeze_header: Freeze the header row
        """
        self.apply_formatting = apply_formatting
        self.auto_filter = auto_filter
        self.freeze_header = freeze_header
    
    def generate(self, 
                emails: List[Dict[str, any]], 
                output_path: str,
                columns: Optional[List[str]] = None,
                sheet_name: str = "Emails") -> str:
        """
        Generate Excel file from email data.
        
        Args:
            emails: List of email dictionaries
            output_path: Output file path
            columns: Column names to include (uses all if None)
            sheet_name: Name of the worksheet
            
        Returns:
            Path to generated file
        """
        if not emails:
            logger.warning("No emails provided for Excel generation")
            return None
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Determine columns
            if columns is None:
                columns = list(emails[0].keys())
            
            # Write headers
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = column
                
                if self.apply_formatting:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.CELL_BORDER
            
            # Write data rows
            for row_idx, email_data in enumerate(emails, start=2):
                for col_idx, column in enumerate(columns, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = email_data.get(column, '')
                    
                    if self.apply_formatting:
                        cell.border = self.CELL_BORDER
                        cell.alignment = Alignment(vertical='center')
            
            # Auto-adjust column widths
            if self.apply_formatting:
                self._auto_adjust_columns(ws, columns)
            
            # Apply auto-filter
            if self.auto_filter:
                ws.auto_filter.ref = ws.dimensions
            
            # Freeze header row
            if self.freeze_header:
                ws.freeze_panes = ws['A2']
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Generated Excel file with {len(emails)} rows at {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate Excel: {str(e)}")
            raise
    
    def generate_simple(self, 
                       emails: List[str], 
                       output_path: str,
                       sheet_name: str = "Emails",
                       include_index: bool = True) -> str:
        """
        Generate simple Excel with just email addresses.
        
        Args:
            emails: List of email addresses
            output_path: Output file path
            sheet_name: Name of the worksheet
            include_index: Include row numbers
            
        Returns:
            Path to generated file
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers
            if include_index:
                headers = ['Index', 'Email']
            else:
                headers = ['Email']
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                
                if self.apply_formatting:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.CELL_BORDER
            
            # Write emails
            for row_idx, email in enumerate(emails, start=2):
                if include_index:
                    ws.cell(row=row_idx, column=1).value = row_idx - 1
                    ws.cell(row=row_idx, column=2).value = email
                else:
                    ws.cell(row=row_idx, column=1).value = email
                
                if self.apply_formatting:
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = self.CELL_BORDER
                        cell.alignment = Alignment(vertical='center')
            
            # Auto-adjust columns
            if self.apply_formatting:
                self._auto_adjust_columns(ws, headers)
            
            # Apply auto-filter
            if self.auto_filter:
                ws.auto_filter.ref = ws.dimensions
            
            # Freeze header
            if self.freeze_header:
                ws.freeze_panes = ws['A2']
            
            # Save
            wb.save(output_path)
            
            logger.info(f"Generated simple Excel with {len(emails)} emails at {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate simple Excel: {str(e)}")
            raise
    
    def _auto_adjust_columns(self, ws, columns: List[str]):
        """
        Auto-adjust column widths based on content.
        
        Args:
            ws: Worksheet object
            columns: List of column names
        """
        for col_idx, column in enumerate(columns, start=1):
            column_letter = get_column_letter(col_idx)
            
            # Get maximum length in column
            max_length = len(str(column))  # Start with header length
            
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Set column width (add padding)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


class EmailFileGenerator:
    """
    High-level generator for creating email output files.
    """
    
    def __init__(self):
        """Initialize file generator."""
        self.csv_generator = CSVGenerator()
        self.excel_generator = ExcelGenerator()
    
    def generate_output(self,
                       emails: List[Dict[str, any]],
                       output_dir: str,
                       filename_prefix: str = "extracted_emails",
                       formats: List[str] = ['csv', 'xlsx']) -> Dict[str, str]:
        """
        Generate output files in multiple formats.
        
        Args:
            emails: List of email dictionaries
            output_dir: Output directory
            filename_prefix: Prefix for filenames
            formats: List of formats to generate ('csv', 'xlsx')
            
        Returns:
            Dictionary mapping format to file path
        """
        if not emails:
            logger.warning("No emails to generate files for")
            return {}
        
        # Create output directory if needed
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate timestamp for unique filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        generated_files = {}
        
        try:
            # Generate CSV
            if 'csv' in formats:
                csv_path = os.path.join(output_dir, f"{filename_prefix}_{timestamp}.csv")
                self.csv_generator.generate(emails, csv_path)
                generated_files['csv'] = csv_path
            
            # Generate Excel
            if 'xlsx' in formats:
                xlsx_path = os.path.join(output_dir, f"{filename_prefix}_{timestamp}.xlsx")
                self.excel_generator.generate(emails, xlsx_path)
                generated_files['xlsx'] = xlsx_path
            
            logger.info(f"Generated {len(generated_files)} output files in {output_dir}")
            return generated_files
            
        except Exception as e:
            logger.error(f"Failed to generate output files: {str(e)}")
            raise
    
    def create_standard_format(self, emails: List[str]) -> List[Dict[str, any]]:
        """
        Convert simple email list to standard format with metadata.
        
        Args:
            emails: List of email addresses
            
        Returns:
            List of email dictionaries
        """
        formatted_emails = []
        
        for idx, email in enumerate(emails, start=1):
            # Split email into username and domain
            if '@' in email:
                username, domain = email.split('@', 1)
            else:
                username, domain = email, ''
            
            formatted_emails.append({
                'index': idx,
                'email': email,
                'username': username,
                'domain': domain,
                'extracted_at': datetime.now().isoformat()
            })
        
        return formatted_emails
    
    def create_detailed_format(self,
                              email_info_list: List[any]) -> List[Dict[str, any]]:
        """
        Convert EmailInfo objects to detailed format.
        
        Args:
            email_info_list: List of EmailInfo objects
            
        Returns:
            List of detailed email dictionaries
        """
        formatted_emails = []
        
        for idx, email_info in enumerate(email_info_list, start=1):
            formatted_emails.append({
                'index': idx,
                'email': email_info.email,
                'username': email_info.username,
                'domain': email_info.domain,
                'confidence_score': round(email_info.confidence_score, 2),
                'is_valid': email_info.is_valid,
                'context': email_info.source_context[:100] if email_info.source_context else '',
                'extracted_at': datetime.now().isoformat()
            })
        
        return formatted_emails


# Convenience functions
def generate_csv(emails: List[str], output_path: str) -> str:
    """
    Convenience function to generate CSV.
    
    Args:
        emails: List of email addresses
        output_path: Output file path
        
    Returns:
        Path to generated file
    """
    generator = CSVGenerator()
    return generator.generate_simple(emails, output_path)


def generate_excel(emails: List[str], output_path: str) -> str:
    """
    Convenience function to generate Excel.
    
    Args:
        emails: List of email addresses
        output_path: Output file path
        
    Returns:
        Path to generated file
    """
    generator = ExcelGenerator()
    return generator.generate_simple(emails, output_path)